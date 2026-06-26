import os
import json
import urllib.request
import base64
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler
import threading
import time
import io
# email via Resend API
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False
    print("[EXCEL] openpyxl nao disponivel")

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "")
ANTHROPIC_KEY = os.environ.get("ANTHROPIC_KEY", "")
TELEGRAM_API = f"https://api.telegram.org/bot{TELEGRAM_TOKEN}"
DATA_FILE = "keyco_data.json"
ADMIN_CHAT_ID = 8601577256  # Chat ID do Filipi
EMAIL_TO = "atendimentokeyco@gmail.com"
RESEND_KEY = "re_ds6QLRf3_9cGRLXjeJM16xDgSGYyvRXjM"  # Resend API key

def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"pagar": [], "receber": [], "orcamentos": [], "counter": {"pag": 0, "cob": 0, "orc": 0}}

def save_data(data):
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def next_id(data, prefix):
    key = {"PAG": "pag", "COB": "cob", "ORC": "orc"}[prefix]
    data["counter"][key] += 1
    return f"{prefix}-{data['counter'][key]:03d}"

def fmt_brl(valor):
    try:
        v = float(valor)
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return f"R$ {valor}"

def days_until(venc_str):
    try:
        d, m, y = venc_str.split("/")
        venc = datetime(int(y), int(m), int(d))
        hoje = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
        return (venc - hoje).days
    except:
        return 999

def status_venc(days):
    if days < 0: return "🔴"
    elif days == 0: return "🔴"
    elif days <= 3: return "🟠"
    elif days <= 7: return "🟡"
    else: return "🟢"

def mes_atual():
    return datetime.now().strftime("%m/%Y")

def is_mes_atual(venc_str):
    try:
        parts = venc_str.split("/")
        return parts[1] == str(datetime.now().month).zfill(2) and parts[2] == str(datetime.now().year)
    except:
        return False

def http_post(url, data):
    body = json.dumps(data).encode()
    req = urllib.request.Request(url, data=body, headers={"Content-Type": "application/json"})
    with urllib.request.urlopen(req, timeout=15) as resp:
        return json.loads(resp.read())

def http_get(url):
    with urllib.request.urlopen(url, timeout=15) as resp:
        return resp.read()

def call_anthropic(payload):
    body = json.dumps(payload).encode()
    req = urllib.request.Request(
        "https://api.anthropic.com/v1/messages",
        data=body,
        headers={"Content-Type": "application/json", "x-api-key": ANTHROPIC_KEY, "anthropic-version": "2023-06-01"}
    )
    with urllib.request.urlopen(req, timeout=30) as resp:
        data = json.loads(resp.read())
        return data.get("content", [{}])[0].get("text", "")

def send_message(chat_id, text, keyboard=None):
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "Markdown"}
    if keyboard:
        payload["reply_markup"] = {"inline_keyboard": keyboard}
    try:
        http_post(f"{TELEGRAM_API}/sendMessage", payload)
    except Exception as e:
        print(f"Erro send: {e}")

def answer_callback(callback_id):
    try:
        http_post(f"{TELEGRAM_API}/answerCallbackQuery", {"callback_query_id": callback_id})
    except:
        pass

# ── KEYBOARDS ──
def kb_menu_principal():
    return [
        [{"text": "💸 A Pagar", "callback_data": "menu_pagar"},
         {"text": "📥 A Receber", "callback_data": "menu_receber"}],
        [{"text": "📋 Orçamentos", "callback_data": "menu_orcamentos"},
         {"text": "📊 Resumo", "callback_data": "acao_resumo"}],
        [{"text": "➕ Nova conta", "callback_data": "acao_nova_pagar"},
         {"text": "📷 Enviar boleto", "callback_data": "acao_foto_boleto"}]
    ]

def kb_pagar():
    return [
        [{"text": "🔴 Vencidas", "callback_data": "pagar_vencidas"},
         {"text": "🔴 Vence hoje", "callback_data": "pagar_hoje"}],
        [{"text": "🟡 Esta semana", "callback_data": "pagar_semana"},
         {"text": "📅 Este mês", "callback_data": "pagar_mes"}],
        [{"text": "📋 Em aberto", "callback_data": "pagar_aberto"},
         {"text": "✅ Pagas", "callback_data": "pagar_pagas"}],
        [{"text": "📊 Todas", "callback_data": "pagar_todas"},
         {"text": "➕ Nova conta", "callback_data": "acao_nova_pagar"}],
        [{"text": "📷 Enviar boleto", "callback_data": "acao_foto_boleto"},
         {"text": "🏠 Menu", "callback_data": "menu_principal"}]
    ]

def kb_receber():
    return [
        [{"text": "🔴 Vencidas", "callback_data": "receber_vencidas"},
         {"text": "🔴 Vence hoje", "callback_data": "receber_hoje"}],
        [{"text": "🟡 Esta semana", "callback_data": "receber_semana"},
         {"text": "📅 Este mês", "callback_data": "receber_mes"}],
        [{"text": "📋 Em aberto", "callback_data": "receber_aberto"},
         {"text": "✅ Recebidas", "callback_data": "receber_recebidas"}],
        [{"text": "📊 Todas", "callback_data": "receber_todas"},
         {"text": "➕ Nova cobrança", "callback_data": "acao_nova_receber"}],
        [{"text": "🏠 Menu", "callback_data": "menu_principal"}]
    ]

def kb_orcamentos():
    return [
        [{"text": "🟡 Pendentes", "callback_data": "orc_pendentes"},
         {"text": "🟢 Aprovados", "callback_data": "orc_aprovados"}],
        [{"text": "🔴 Recusados", "callback_data": "orc_recusados"},
         {"text": "🔵 Enviados", "callback_data": "orc_enviados"}],
        [{"text": "📊 Todos", "callback_data": "orc_todos"},
         {"text": "➕ Novo orçamento", "callback_data": "acao_novo_orc"}],
        [{"text": "🏠 Menu", "callback_data": "menu_principal"}]
    ]

def kb_voltar_pagar():
    return [[{"text": "◀ Voltar", "callback_data": "menu_pagar"}, {"text": "🏠 Menu", "callback_data": "menu_principal"}]]

def kb_voltar_receber():
    return [[{"text": "◀ Voltar", "callback_data": "menu_receber"}, {"text": "🏠 Menu", "callback_data": "menu_principal"}]]

def kb_voltar_orc():
    return [[{"text": "◀ Voltar", "callback_data": "menu_orcamentos"}, {"text": "🏠 Menu", "callback_data": "menu_principal"}]]

# ── LISTAS ──
def lista_pagar(items, titulo):
    if not items:
        return f"{titulo}\n\n✅ Nenhuma conta nesta categoria."
    linhas = [f"{titulo}\n"]
    total = 0
    for p in sorted(items, key=lambda x: days_until(x["vencimento"]) if x["status"]!="pago" else 999):
        if p["status"] == "pago":
            linhas.append(f"✅ *{p['id']}* — {p['fornecedor']}\n💰 {fmt_brl(p['valor'])} · pago {p.get('data_pagamento','—')}\n")
        else:
            d = days_until(p["vencimento"])
            st = status_venc(d)
            dias_txt = f"({abs(d)}d atrás)" if d < 0 else f"(em {d}d)" if d > 0 else "(HOJE)"
            linhas.append(f"{st} *{p['id']}* — {p['fornecedor']}\n💰 {fmt_brl(p['valor'])} · {p['vencimento']} {dias_txt}\n")
            total += float(p["valor"])
    if total > 0:
        linhas.append(f"*Total: {fmt_brl(total)}*")
    return "\n".join(linhas)

def lista_receber(items, titulo):
    if not items:
        return f"{titulo}\n\n✅ Nenhuma cobrança nesta categoria."
    linhas = [f"{titulo}\n"]
    total = 0
    for r in sorted(items, key=lambda x: days_until(x["vencimento"]) if x["status"]!="pago" else 999):
        if r["status"] == "pago":
            linhas.append(f"✅ *{r['id']}* — {r['cliente']}\n💰 {fmt_brl(r['valor'])} · recebido {r.get('data_pagamento','—')}\n")
        else:
            d = days_until(r["vencimento"])
            st = status_venc(d)
            dias_txt = f"({abs(d)}d atrás)" if d < 0 else f"(em {d}d)" if d > 0 else "(HOJE)"
            linhas.append(f"{st} *{r['id']}* — {r['cliente']}\n💰 {fmt_brl(r['valor'])} · {r['vencimento']} {dias_txt}\n")
            total += float(r["valor"])
    if total > 0:
        linhas.append(f"*Total: {fmt_brl(total)}*")
    return "\n".join(linhas)

def lista_orcamentos(items, titulo):
    if not items:
        return f"{titulo}\n\nNenhum orçamento nesta categoria."
    emoji = {"pendente":"🟡","aprovado":"🟢","recusado":"🔴","enviado":"🔵"}
    linhas = [f"{titulo}\n"]
    total = sum(float(o["valor"]) for o in items)
    for o in items:
        em = emoji.get(o["status"],"⚪")
        linhas.append(f"{em} *{o['id']}* — {o['cliente']}\n📋 {o['descricao']}\n💰 {fmt_brl(o['valor'])} · {o['data']}\n")
    linhas.append(f"*Total: {fmt_brl(total)}*")
    return "\n".join(linhas)

# ── CALLBACKS ──
def handle_callback(chat_id, data_cb):
    db = load_data()
    hoje = datetime.now()

    # MENUS
    if data_cb == "menu_principal":
        send_message(chat_id, "🏢 *Keycomerce — Financeiro*\n\nEscolha uma opção:", kb_menu_principal())
        return

    if data_cb == "menu_pagar":
        ab = len([p for p in db["pagar"] if p["status"]=="aberto"])
        venc = len([p for p in db["pagar"] if p["status"]=="aberto" and days_until(p["vencimento"])<0])
        send_message(chat_id, f"💸 *Contas a Pagar*\n\n📋 Em aberto: {ab} | 🔴 Vencidas: {venc}", kb_pagar())
        return

    if data_cb == "menu_receber":
        ab = len([r for r in db["receber"] if r["status"]=="aberto"])
        venc = len([r for r in db["receber"] if r["status"]=="aberto" and days_until(r["vencimento"])<0])
        send_message(chat_id, f"📥 *Contas a Receber*\n\n📋 Em aberto: {ab} | 🔴 Vencidas: {venc}", kb_receber())
        return

    if data_cb == "menu_orcamentos":
        pend = len([o for o in db["orcamentos"] if o["status"]=="pendente"])
        send_message(chat_id, f"📋 *Orçamentos*\n\n🟡 Pendentes: {pend}", kb_orcamentos())
        return

    # PAGAR
    if data_cb == "pagar_vencidas":
        items = [p for p in db["pagar"] if p["status"]=="aberto" and days_until(p["vencimento"])<0]
        send_message(chat_id, lista_pagar(items, "🔴 *Contas vencidas*"), kb_voltar_pagar())
        return

    if data_cb == "pagar_hoje":
        items = [p for p in db["pagar"] if p["status"]=="aberto" and days_until(p["vencimento"])==0]
        send_message(chat_id, lista_pagar(items, "🔴 *Vence hoje*"), kb_voltar_pagar())
        return

    if data_cb == "pagar_semana":
        items = [p for p in db["pagar"] if p["status"]=="aberto" and 0<=days_until(p["vencimento"])<=7]
        send_message(chat_id, lista_pagar(items, "🟡 *Vence esta semana*"), kb_voltar_pagar())
        return

    if data_cb == "pagar_mes":
        items = [p for p in db["pagar"] if p["status"]=="aberto" and is_mes_atual(p["vencimento"])]
        send_message(chat_id, lista_pagar(items, f"📅 *A pagar — {mes_atual()}*"), kb_voltar_pagar())
        return

    if data_cb == "pagar_aberto":
        items = [p for p in db["pagar"] if p["status"]=="aberto"]
        send_message(chat_id, lista_pagar(items, "📋 *Todas em aberto*"), kb_voltar_pagar())
        return

    if data_cb == "pagar_pagas":
        items = [p for p in db["pagar"] if p["status"]=="pago"]
        send_message(chat_id, lista_pagar(items, "✅ *Contas pagas*"), kb_voltar_pagar())
        return

    if data_cb == "pagar_todas":
        items = db["pagar"]
        send_message(chat_id, lista_pagar(items, "📊 *Todas as contas a pagar*"), kb_voltar_pagar())
        return

    # RECEBER
    if data_cb == "receber_vencidas":
        items = [r for r in db["receber"] if r["status"]=="aberto" and days_until(r["vencimento"])<0]
        send_message(chat_id, lista_receber(items, "🔴 *Cobranças vencidas*"), kb_voltar_receber())
        return

    if data_cb == "receber_hoje":
        items = [r for r in db["receber"] if r["status"]=="aberto" and days_until(r["vencimento"])==0]
        send_message(chat_id, lista_receber(items, "🔴 *Vence hoje*"), kb_voltar_receber())
        return

    if data_cb == "receber_semana":
        items = [r for r in db["receber"] if r["status"]=="aberto" and 0<=days_until(r["vencimento"])<=7]
        send_message(chat_id, lista_receber(items, "🟡 *Vence esta semana*"), kb_voltar_receber())
        return

    if data_cb == "receber_mes":
        items = [r for r in db["receber"] if r["status"]=="aberto" and is_mes_atual(r["vencimento"])]
        send_message(chat_id, lista_receber(items, f"📅 *A receber — {mes_atual()}*"), kb_voltar_receber())
        return

    if data_cb == "receber_aberto":
        items = [r for r in db["receber"] if r["status"]=="aberto"]
        send_message(chat_id, lista_receber(items, "📋 *Todas em aberto*"), kb_voltar_receber())
        return

    if data_cb == "receber_recebidas":
        items = [r for r in db["receber"] if r["status"]=="pago"]
        send_message(chat_id, lista_receber(items, "✅ *Cobranças recebidas*"), kb_voltar_receber())
        return

    if data_cb == "receber_todas":
        items = db["receber"]
        send_message(chat_id, lista_receber(items, "📊 *Todas as cobranças*"), kb_voltar_receber())
        return

    # ORÇAMENTOS
    if data_cb == "orc_pendentes":
        items = [o for o in db["orcamentos"] if o["status"]=="pendente"]
        send_message(chat_id, lista_orcamentos(items, "🟡 *Orçamentos pendentes*"), kb_voltar_orc())
        return

    if data_cb == "orc_aprovados":
        items = [o for o in db["orcamentos"] if o["status"]=="aprovado"]
        send_message(chat_id, lista_orcamentos(items, "🟢 *Orçamentos aprovados*"), kb_voltar_orc())
        return

    if data_cb == "orc_recusados":
        items = [o for o in db["orcamentos"] if o["status"]=="recusado"]
        send_message(chat_id, lista_orcamentos(items, "🔴 *Orçamentos recusados*"), kb_voltar_orc())
        return

    if data_cb == "orc_enviados":
        items = [o for o in db["orcamentos"] if o["status"]=="enviado"]
        send_message(chat_id, lista_orcamentos(items, "🔵 *Orçamentos enviados*"), kb_voltar_orc())
        return

    if data_cb == "orc_todos":
        items = db["orcamentos"]
        send_message(chat_id, lista_orcamentos(items, "📊 *Todos os orçamentos*"), kb_voltar_orc())
        return

    # AÇÕES
    if data_cb == "acao_resumo":
        pag = [p for p in db["pagar"] if p["status"]=="aberto"]
        rec = [r for r in db["receber"] if r["status"]=="aberto"]
        orc_pend = [o for o in db["orcamentos"] if o["status"]=="pendente"]
        venc_p = [p for p in pag if days_until(p["vencimento"])<0]
        venc_r = [r for r in rec if days_until(r["vencimento"])<0]
        hoje_p = [p for p in pag if days_until(p["vencimento"])==0]
        semana_p = [p for p in pag if 1<=days_until(p["vencimento"])<=7]
        total_pago = sum(float(p["valor"]) for p in db["pagar"] if p["status"]=="pago")
        total_recebido = sum(float(r["valor"]) for r in db["receber"] if r["status"]=="pago")
        msg = f"""📊 *Resumo Financeiro — Keyco*
📅 {datetime.now().strftime('%d/%m/%Y')}

💸 *A PAGAR*
Em aberto: {fmt_brl(sum(float(p['valor']) for p in pag))}
🔴 Vencidas: {len(venc_p)} conta(s)
🔴 Vence hoje: {len(hoje_p)} conta(s)
🟡 Esta semana: {len(semana_p)} conta(s)
✅ Total pago no histórico: {fmt_brl(total_pago)}

📥 *A RECEBER*
Em aberto: {fmt_brl(sum(float(r['valor']) for r in rec))}
🔴 Vencidas: {len(venc_r)} cobrança(s)
✅ Total recebido no histórico: {fmt_brl(total_recebido)}

📋 *ORÇAMENTOS*
🟡 Pendentes: {len(orc_pend)}
Total em orçamentos pendentes: {fmt_brl(sum(float(o['valor']) for o in orc_pend))}"""
        send_message(chat_id, msg, [[{"text": "🏠 Menu", "callback_data": "menu_principal"}]])
        return

    if data_cb == "acao_nova_pagar":
        user_states[chat_id] = {"step": "nova_pagar_manual", "tipo_lancamento": "pagar"}
        send_message(chat_id, "💸 *Nova conta a pagar*\n\nMande 📷 foto ou PDF do boleto\n\nOu digite:\n*Fornecedor, Valor, dd/mm/aaaa*\nEx: `Udinese perfis, 3200, 25/07/2026`")
        return

    if data_cb == "acao_nova_receber":
        user_states[chat_id] = {"step": "nova_receber_manual", "tipo_lancamento": "receber"}
        send_message(chat_id, "📥 *Nova cobrança*\n\nDigite:\n*Cliente, Valor, dd/mm/aaaa*\nEx: `Esquadrias João, 1640, 25/07/2026`")
        return

    if data_cb == "acao_novo_orc":
        user_states[chat_id] = {"step": "novo_orc_manual"}
        send_message(chat_id, "📋 *Novo orçamento*\n\nDigite:\n*Cliente, Descrição, Valor*\nEx: `Esquadrias João, 10 telas Udinese branco, 890`")
        return

    if data_cb == "acao_foto_boleto":
        user_states[chat_id] = {"step": "aguardando_foto", "tipo_lancamento": "pagar"}
        send_message(chat_id, "📷 Mande a foto ou PDF do boleto agora:")
        return

# ── STATE ──
user_states = {}

def handle_message(chat_id, text=None, photo=None, document=None):
    db = load_data()
    state = user_states.get(chat_id, {})

    # FOTO/DOC
    if photo or document:
        file_id = photo[-1]["file_id"] if photo else document["file_id"]
        send_message(chat_id, "⏳ Lendo o documento com IA...")
        resultado = read_boleto_image(file_id)
        if resultado:
            user_states[chat_id] = {"step": "confirmar_boleto", "temp": resultado, "tipo": state.get("tipo_lancamento", "pagar")}
            tipo = "pagar" if state.get("tipo_lancamento", "pagar") == "pagar" else "receber"
            kb = [[{"text": "✅ Confirmar", "callback_data": f"confirmar_boleto_{tipo}"}, {"text": "❌ Cancelar", "callback_data": "menu_principal"}]]
            send_message(chat_id, f"✅ *Dados extraídos:*\n\n🏢 {resultado.get('fornecedor','—')}\n💰 {fmt_brl(resultado.get('valor',0))}\n📅 {resultado.get('vencimento','—')}\n\nConfirmar como conta a {tipo}?", kb)
        else:
            send_message(chat_id, "❌ Não consegui ler. Digite manualmente após clicar em Nova conta.", [[{"text": "🏠 Menu", "callback_data": "menu_principal"}]])
        return

    if not text:
        return
    tl = text.lower().strip()

    # ESCAPE DE QUALQUER ESTADO
    if tl in ["menu", "/menu", "/start", "oi", "olá", "ola", "cancelar", "cancel", "sair", "voltar", "escape", "inicio", "início"]:
        user_states[chat_id] = {}
        send_message(chat_id, "🏢 *Keycomerce — Financeiro*\n\nEscolha uma opção:", kb_menu_principal())
        return

    # ESTADOS DE CONVERSA
    if state.get("step") == "nova_pagar_manual":
        parts = [p.strip() for p in text.split(",")]
        if len(parts) >= 3:
            try:
                val = float(parts[1].replace("R$","").replace(".","").replace(",",".").strip())
                item = {"id": next_id(db,"PAG"), "fornecedor": parts[0], "valor": val, "vencimento": parts[2].strip(), "status": "aberto", "tipo": "Manual"}
                db["pagar"].append(item)
                save_data(db)
                user_states[chat_id] = {}
                send_message(chat_id, f"✅ *{item['id']}* lançado!\n🏢 {item['fornecedor']}\n💰 {fmt_brl(item['valor'])}\n📅 {item['vencimento']}", [[{"text": "💸 Ver contas", "callback_data": "menu_pagar"}, {"text": "🏠 Menu", "callback_data": "menu_principal"}]])
            except:
                send_message(chat_id, "❌ Valor inválido. Use: *Fornecedor, 3200, 25/07/2026*")
        else:
            send_message(chat_id, "❌ Formato: *Fornecedor, Valor, dd/mm/aaaa*\nEx: `Udinese, 3200, 25/07/2026`")
        return

    if state.get("step") == "nova_receber_manual":
        parts = [p.strip() for p in text.split(",")]
        if len(parts) >= 3:
            try:
                val = float(parts[1].replace("R$","").replace(".","").replace(",",".").strip())
                item = {"id": next_id(db,"COB"), "cliente": parts[0], "valor": val, "vencimento": parts[2].strip(), "status": "aberto", "nf": ""}
                db["receber"].append(item)
                save_data(db)
                user_states[chat_id] = {}
                send_message(chat_id, f"✅ *{item['id']}* lançado!\n👤 {item['cliente']}\n💰 {fmt_brl(item['valor'])}\n📅 {item['vencimento']}", [[{"text": "📥 Ver cobranças", "callback_data": "menu_receber"}, {"text": "🏠 Menu", "callback_data": "menu_principal"}]])
            except:
                send_message(chat_id, "❌ Valor inválido.")
        else:
            send_message(chat_id, "❌ Formato: *Cliente, Valor, dd/mm/aaaa*")
        return

    if state.get("step") == "novo_orc_manual":
        parts = [p.strip() for p in text.split(",")]
        if len(parts) >= 3:
            try:
                val = float(parts[2].replace("R$","").replace(".","").replace(",",".").strip())
                item = {"id": next_id(db,"ORC"), "cliente": parts[0], "descricao": parts[1], "valor": val, "status": "pendente", "data": datetime.now().strftime("%d/%m/%Y")}
                db["orcamentos"].append(item)
                save_data(db)
                user_states[chat_id] = {}
                send_message(chat_id, f"✅ *{item['id']}* criado!\n👤 {item['cliente']}\n📋 {item['descricao']}\n💰 {fmt_brl(item['valor'])}", [[{"text": "📋 Ver orçamentos", "callback_data": "menu_orcamentos"}, {"text": "🏠 Menu", "callback_data": "menu_principal"}]])
            except:
                send_message(chat_id, "❌ Valor inválido.")
        else:
            send_message(chat_id, "❌ Formato: *Cliente, Descrição, Valor*")
        return

    if state.get("step") == "confirmar_boleto":
        if "sim" in tl:
            temp = state["temp"]
            tipo = state.get("tipo", "pagar")
            if tipo == "pagar":
                item = {"id": next_id(db,"PAG"), "fornecedor": temp.get("fornecedor","Sem nome"), "valor": float(temp.get("valor",0)), "vencimento": temp.get("vencimento",""), "status": "aberto", "tipo": "Boleto"}
                db["pagar"].append(item)
                save_data(db)
                user_states[chat_id] = {}
                send_message(chat_id, f"✅ *{item['id']}* lançado!\n🏢 {item['fornecedor']}\n💰 {fmt_brl(item['valor'])}\n📅 {item['vencimento']}", [[{"text": "💸 Ver contas", "callback_data": "menu_pagar"}, {"text": "🏠 Menu", "callback_data": "menu_principal"}]])
            else:
                item = {"id": next_id(db,"COB"), "cliente": temp.get("fornecedor","Sem nome"), "valor": float(temp.get("valor",0)), "vencimento": temp.get("vencimento",""), "status": "aberto", "nf": ""}
                db["receber"].append(item)
                save_data(db)
                user_states[chat_id] = {}
                send_message(chat_id, f"✅ *{item['id']}* lançado!", [[{"text": "🏠 Menu", "callback_data": "menu_principal"}]])
        else:
            user_states[chat_id] = {}
            send_message(chat_id, "❌ Cancelado.", [[{"text": "🏠 Menu", "callback_data": "menu_principal"}]])
        return

    # PLANILHA / EMAIL
    if tl in ["email", "mandar email", "enviar email", "planilha email", "e-mail", "mandar e-mail"]:
        send_message(chat_id, "⏳ Gerando e enviando por e-mail...")
        def _enviar():
            excel_bytes = gerar_excel()
            if excel_bytes:
                ok = enviar_email_planilha(excel_bytes, datetime.now().strftime("%d/%m/%Y"))
                send_message(chat_id, "✅ Planilha enviada para atendimentokeyco@gmail.com!" if ok else "❌ Erro ao enviar e-mail.")
            else:
                send_message(chat_id, "❌ Erro ao gerar planilha.")
        threading.Thread(target=_enviar).start()
        return

    # PLANILHA
    if tl in ["planilha", "gerar planilha", "exportar", "excel", "relatorio", "relatório"]:
        send_message(chat_id, "⏳ Gerando planilha...")
        threading.Thread(target=enviar_planilha, args=(chat_id, "📊 Planilha gerada agora")).start()
        return

    # MARCAR PAGO POR TEXTO
    if tl.startswith("paguei") or tl.startswith("pago "):
        nome = tl.replace("paguei","").replace("pago","").strip()
        encontrados = [p for p in db["pagar"] if p["status"]=="aberto" and nome and nome in p["fornecedor"].lower()]
        if not encontrados:
            send_message(chat_id, f"❌ Nenhuma conta aberta com *{nome}*.", [[{"text": "💸 Ver contas", "callback_data": "menu_pagar"}]])
        elif len(encontrados) == 1:
            encontrados[0]["status"] = "pago"
            encontrados[0]["data_pagamento"] = datetime.now().strftime("%d/%m/%Y")
            save_data(db)
            send_message(chat_id, f"✅ *{encontrados[0]['id']}* marcado como pago!\n🏢 {encontrados[0]['fornecedor']} — {fmt_brl(encontrados[0]['valor'])}", [[{"text": "💸 Ver contas", "callback_data": "menu_pagar"}, {"text": "🏠 Menu", "callback_data": "menu_principal"}]])
        else:
            linhas = [f"Encontrei {len(encontrados)} contas:\n"]
            for p in encontrados:
                linhas.append(f"• *{p['id']}* — {p['fornecedor']} · {fmt_brl(p['valor'])} · {p['vencimento']}")
            linhas.append("\nDigite o ID para marcar. Ex: *PAG-001*")
            send_message(chat_id, "\n".join(linhas))
        return

    if tl.startswith("recebido ") or tl.startswith("recebi "):
        nome = tl.replace("recebido","").replace("recebi","").strip()
        encontrados = [r for r in db["receber"] if r["status"]=="aberto" and nome and nome in r["cliente"].lower()]
        if not encontrados:
            send_message(chat_id, f"❌ Nenhuma cobrança aberta com *{nome}*.", [[{"text": "📥 Ver cobranças", "callback_data": "menu_receber"}]])
        elif len(encontrados) == 1:
            encontrados[0]["status"] = "pago"
            encontrados[0]["data_pagamento"] = datetime.now().strftime("%d/%m/%Y")
            save_data(db)
            send_message(chat_id, f"✅ *{encontrados[0]['id']}* marcado como recebido!\n👤 {encontrados[0]['cliente']} — {fmt_brl(encontrados[0]['valor'])}", [[{"text": "📥 Ver cobranças", "callback_data": "menu_receber"}, {"text": "🏠 Menu", "callback_data": "menu_principal"}]])
        else:
            linhas = [f"Encontrei {len(encontrados)} cobranças:"]
            for r in encontrados:
                linhas.append(f"• *{r['id']}* — {r['cliente']} · {fmt_brl(r['valor'])}")
            linhas.append("\nDigite o ID. Ex: *COB-001*")
            send_message(chat_id, "\n".join(linhas))
        return

    # ID DIRETO
    if text.upper().startswith("PAG-"):
        item = next((p for p in db["pagar"] if p["id"]==text.upper().strip()), None)
        if item and item["status"]=="aberto":
            item["status"] = "pago"
            item["data_pagamento"] = datetime.now().strftime("%d/%m/%Y")
            save_data(db)
            send_message(chat_id, f"✅ *{item['id']}* marcado como pago!", [[{"text": "💸 Ver contas", "callback_data": "menu_pagar"}]])
        return

    if text.upper().startswith("COB-"):
        item = next((r for r in db["receber"] if r["id"]==text.upper().strip()), None)
        if item and item["status"]=="aberto":
            item["status"] = "pago"
            item["data_pagamento"] = datetime.now().strftime("%d/%m/%Y")
            save_data(db)
            send_message(chat_id, f"✅ *{item['id']}* marcado como recebido!", [[{"text": "📥 Ver cobranças", "callback_data": "menu_receber"}]])
        return

    if text.upper().startswith("ORC-"):
        orc_id = text.upper().strip()
        orc = next((o for o in db["orcamentos"] if o["id"]==orc_id), None)
        if orc:
            kb = [[{"text": "✅ Aprovar", "callback_data": f"orc_aprovar_{orc_id}"}, {"text": "❌ Recusar", "callback_data": f"orc_recusar_{orc_id}"}], [{"text": "🏠 Menu", "callback_data": "menu_principal"}]]
            send_message(chat_id, f"📋 *{orc['id']}*\n👤 {orc['cliente']}\n📋 {orc['descricao']}\n💰 {fmt_brl(orc['valor'])}\nStatus: {orc['status'].capitalize()}", kb)
        return

    # APROVAR/RECUSAR POR TEXTO
    if tl.startswith("aprovar "):
        orc_id = text.upper().replace("APROVAR","").strip()
        orc = next((o for o in db["orcamentos"] if o["id"]==orc_id), None)
        if orc:
            orc["status"] = "aprovado"
            save_data(db)
            send_message(chat_id, f"✅ *{orc['id']}* aprovado!", [[{"text": "📋 Orçamentos", "callback_data": "menu_orcamentos"}]])
        return

    if tl.startswith("recusar "):
        orc_id = text.upper().replace("RECUSAR","").strip()
        orc = next((o for o in db["orcamentos"] if o["id"]==orc_id), None)
        if orc:
            orc["status"] = "recusado"
            save_data(db)
            send_message(chat_id, f"❌ *{orc['id']}* recusado.", [[{"text": "📋 Orçamentos", "callback_data": "menu_orcamentos"}]])
        return

    # START / MENU
    if tl in ["/start", "/menu", "menu", "oi", "olá", "ola", "inicio", "início"]:
        send_message(chat_id, "🏢 *Keycomerce — Financeiro*\n\nBom dia! Escolha uma opção:", kb_menu_principal())
        return

    # IA FALLBACK
    if ANTHROPIC_KEY:
        try:
            system = """Interprete e retorne JSON: {"intencao": "...", "dados": {"nome":"...","id":"..."}}
Intenções: menu, marcar_pago, marcar_recebido, nova_conta_pagar, nova_conta_receber, novo_orcamento, resumo
Ex: "paguei Udinese" → {"intencao":"marcar_pago","dados":{"nome":"Udinese"}}"""
            result = call_anthropic({"model": "claude-sonnet-4-6", "max_tokens": 150, "system": system, "messages": [{"role": "user", "content": text}]})
            result = result.replace("```json","").replace("```","").strip()
            parsed = json.loads(result)
            intencao = parsed.get("intencao","menu")
            dados = parsed.get("dados",{})
            if intencao == "marcar_pago" and dados.get("nome"):
                handle_message(chat_id, text=f"paguei {dados['nome']}")
                return
            if intencao == "marcar_recebido" and dados.get("nome"):
                handle_message(chat_id, text=f"recebido {dados['nome']}")
                return
            if intencao == "nova_conta_pagar":
                handle_callback(chat_id, "acao_nova_pagar")
                return
            if intencao == "nova_conta_receber":
                handle_callback(chat_id, "acao_nova_receber")
                return
            if intencao == "resumo":
                handle_callback(chat_id, "acao_resumo")
                return
        except:
            pass

    send_message(chat_id, "🤔 Não entendi. Use o menu:", kb_menu_principal())

def read_boleto_image(file_id):
    print(f"[BOLETO] Iniciando leitura file_id={file_id}")
    if not ANTHROPIC_KEY:
        print("[BOLETO] Sem ANTHROPIC_KEY")
        return None
    try:
        print("[BOLETO] Buscando file_path no Telegram...")
        file_info = http_get(f"{TELEGRAM_API}/getFile?file_id={file_id}")
        file_data = json.loads(file_info)
        file_path = file_data.get("result", {}).get("file_path", "")
        print(f"[BOLETO] file_path={file_path}")
        if not file_path:
            print("[BOLETO] file_path vazio")
            return None
        print("[BOLETO] Baixando arquivo...")
        file_bytes = http_get(f"https://api.telegram.org/file/bot{TELEGRAM_TOKEN}/{file_path}")
        print(f"[BOLETO] Arquivo: {len(file_bytes)} bytes")
        b64 = base64.b64encode(file_bytes).decode()
        media_type = "image/jpeg"
        if file_path.endswith(".png"): media_type = "image/png"
        elif file_path.endswith(".pdf"): media_type = "application/pdf"
        print(f"[BOLETO] media_type={media_type}, enviando para Claude...")
        prompt = 'Este e um boleto bancario brasileiro. Extraia: fornecedor/beneficiario, valor total, data de vencimento. Responda SOMENTE JSON: {"fornecedor":"...","valor":0.00,"vencimento":"dd/mm/aaaa"}'
        if media_type == "application/pdf":
            content = [{"type": "document", "source": {"type": "base64", "media_type": media_type, "data": b64}}, {"type": "text", "text": prompt}]
        else:
            content = [{"type": "image", "source": {"type": "base64", "media_type": media_type, "data": b64}}, {"type": "text", "text": prompt}]
        result = call_anthropic({"model": "claude-sonnet-4-6", "max_tokens": 300, "messages": [{"role": "user", "content": content}]})
        print(f"[BOLETO] Claude respondeu: {result}")
        cleaned = result.replace("```json","").replace("```","").strip()
        parsed = json.loads(cleaned)
        print(f"[BOLETO] Parsed OK: {parsed}")
        return parsed
    except Exception as e:
        print(f"[BOLETO] ERRO: {type(e).__name__}: {e}")
        return None

# ── EXCEL ──
def gerar_excel():
    if not EXCEL_OK:
        return None
    try:
        from openpyxl.utils import get_column_letter
        db = load_data()
        hoje = datetime.now()
        data_str = hoje.strftime("%d/%m/%Y")
        hora_str = hoje.strftime("%H:%M")

        COR_AZUL="1A4F8A"; COR_AZUL_CLARO="E8F0FA"; COR_VERDE="27AE60"; COR_VERDE_CLARO="D5F5E3"
        COR_VERMELHO="C0392B"; COR_VERMELHO_CLARO="FADBD8"; COR_LARANJA="E67E22"; COR_LARANJA_CLARO="FDEBD0"
        COR_CINZA="7F8C8D"; COR_CINZA_CLARO="F2F3F4"; COR_AMARELO_CLARO="FEF9E7"; COR_BRANCO="FFFFFF"
        COR_PRETO="1A1A1A"; COR_HEADER_DARK="1A2940"

        def fl(cor): return PatternFill("solid", fgColor=cor)
        def fn(bold=False,size=10,cor="1A1A1A",italic=False): return Font(bold=bold,size=size,color=cor,italic=italic,name="Calibri")
        def bd():
            s=Side(style="thin",color="CCCCCC")
            return Border(left=s,right=s,top=s,bottom=s)
        def al(h="left",v="center",wrap=False): return Alignment(horizontal=h,vertical=v,wrap_text=wrap)

        def header_row(ws, headers, row, bg=COR_HEADER_DARK, fg=COR_BRANCO):
            ws.append(headers)
            for cell in ws[row]:
                cell.font=fn(bold=True,size=10,cor=fg); cell.fill=fl(bg)
                cell.alignment=al("center"); cell.border=bd()

        def titulo_sheet(ws, titulo, subtitulo, cor):
            ws.row_dimensions[1].height=40; ws.row_dimensions[2].height=20; ws.row_dimensions[3].height=8
            c1=ws.cell(row=1,column=1,value=titulo)
            c1.font=Font(bold=True,size=16,color=COR_BRANCO,name="Calibri")
            c1.fill=fl(cor); c1.alignment=al("left","center")
            c2=ws.cell(row=2,column=1,value=subtitulo)
            c2.font=fn(size=10,cor="999999",italic=True)
            c2.fill=fl(COR_CINZA_CLARO); c2.alignment=al("left","center")

        def merge_titulo(ws, cols, cor):
            for col in range(1,cols+1):
                ws.cell(row=1,column=col).fill=fl(cor)
                ws.cell(row=2,column=col).fill=fl(COR_CINZA_CLARO)
                ws.cell(row=3,column=col).fill=fl(COR_BRANCO)
            ws.merge_cells(f"A1:{get_column_letter(cols)}1")
            ws.merge_cells(f"A2:{get_column_letter(cols)}2")
            ws.merge_cells(f"A3:{get_column_letter(cols)}3")

        def fmt_v(v):
            try: return float(v)
            except: return 0

        pag_ab=[p for p in db["pagar"] if p["status"]=="aberto"]
        pag_pg=[p for p in db["pagar"] if p["status"]=="pago"]
        pag_venc=[p for p in pag_ab if days_until(p["vencimento"])<0]
        pag_hoje=[p for p in pag_ab if days_until(p["vencimento"])==0]
        pag_sem=[p for p in pag_ab if 1<=days_until(p["vencimento"])<=7]
        rec_ab=[r for r in db["receber"] if r["status"]=="aberto"]
        rec_pg=[r for r in db["receber"] if r["status"]=="pago"]
        rec_venc=[r for r in rec_ab if days_until(r["vencimento"])<0]
        orc_pend=[o for o in db["orcamentos"] if o["status"]=="pendente"]
        orc_aprov=[o for o in db["orcamentos"] if o["status"]=="aprovado"]
        orc_rec=[o for o in db["orcamentos"] if o["status"]=="recusado"]

        total_pagar=sum(fmt_v(p["valor"]) for p in pag_ab)
        total_receber=sum(fmt_v(r["valor"]) for r in rec_ab)
        total_recebido=sum(fmt_v(r["valor"]) for r in rec_pg)
        total_orc_pend=sum(fmt_v(o["valor"]) for o in orc_pend)

        wb = openpyxl.Workbook()

        # ── RESUMO ──
        ws=wb.active; ws.title="📊 Resumo"; ws.sheet_view.showGridLines=False
        ws.row_dimensions[1].height=50; ws.row_dimensions[2].height=22; ws.row_dimensions[3].height=12
        ws.merge_cells("A1:H1")
        c=ws.cell(row=1,column=1,value="  KEYCOMERCE — RELATÓRIO FINANCEIRO")
        c.font=Font(bold=True,size=20,color=COR_BRANCO,name="Calibri"); c.fill=fl(COR_AZUL); c.alignment=al("left","center")
        ws.merge_cells("A2:H2")
        c2=ws.cell(row=2,column=1,value=f"  Gerado em {data_str} às {hora_str}  |  Santa Cruz do Sul — RS")
        c2.font=fn(size=10,cor="555555",italic=True); c2.fill=fl("F0F4FA"); c2.alignment=al("left","center")
        ws.merge_cells("A3:H3"); ws.cell(row=3,column=1).fill=fl(COR_BRANCO)

        def big_card(ws,row,col1,col2,titulo,valor,subtitulo,bg,fg_titulo,fg_valor):
            ws.row_dimensions[row].height=18; ws.row_dimensions[row+1].height=32; ws.row_dimensions[row+2].height=16
            c1=ws.cell(row=row,column=col1,value=titulo.upper())
            c2=ws.cell(row=row+1,column=col1,value=valor)
            c3=ws.cell(row=row+2,column=col1,value=subtitulo)
            ws.merge_cells(start_row=row,start_column=col1,end_row=row,end_column=col2)
            ws.merge_cells(start_row=row+1,start_column=col1,end_row=row+1,end_column=col2)
            ws.merge_cells(start_row=row+2,start_column=col1,end_row=row+2,end_column=col2)
            for r in [row,row+1,row+2]:
                for c in range(col1,col2+1):
                    cell=ws.cell(row=r,column=c); cell.fill=fl(bg); cell.border=bd()
            c1.font=fn(bold=True,size=9,cor=fg_titulo); c1.alignment=al("center","center")
            c2.font=Font(bold=True,size=18,color=fg_valor,name="Calibri"); c2.alignment=al("center","center"); c2.number_format="R$ #,##0.00"
            c3.font=fn(size=8,cor="888888",italic=True); c3.alignment=al("center","center")

        big_card(ws,4,1,2,"💸 Total a Pagar",total_pagar,f"{len(pag_ab)} conta(s) em aberto","FFF5F5","C0392B","C0392B")
        big_card(ws,4,3,4,"📥 Total a Receber",total_receber,f"{len(rec_ab)} cobrança(s) em aberto","F0FFF4","27AE60","27AE60")
        big_card(ws,4,5,6,"✅ Total Recebido",total_recebido,f"{len(rec_pg)} cobrança(s) recebida(s)","F0F8FF","1A4F8A","1A4F8A")
        big_card(ws,4,7,8,"📋 Orç. Pendentes",total_orc_pend,f"{len(orc_pend)} orçamento(s) aguardando","FFFDF0","E67E22","E67E22")

        ws.row_dimensions[7].height=10
        for c in range(1,9): ws.cell(row=7,column=c).fill=fl(COR_BRANCO)

        row=8; ws.row_dimensions[row].height=22
        ws.merge_cells(f"A{row}:H{row}")
        c=ws.cell(row=row,column=1,value="⚠  ALERTAS E PENDÊNCIAS")
        c.font=fn(bold=True,size=11,cor=COR_BRANCO); c.fill=fl(COR_AZUL); c.alignment=al("left","center")

        alertas=[]
        if pag_venc: alertas.append(("🔴 VENCIDO",f"{len(pag_venc)} conta(s) vencida(s) — Total: {fmt_brl(sum(fmt_v(p["valor"]) for p in pag_venc))}",COR_VERMELHO_CLARO,COR_VERMELHO))
        if pag_hoje: alertas.append(("🔴 HOJE",f"{len(pag_hoje)} conta(s) vence(m) hoje — Total: {fmt_brl(sum(fmt_v(p["valor"]) for p in pag_hoje))}",COR_LARANJA_CLARO,COR_LARANJA))
        if pag_sem: alertas.append(("🟡 ESTA SEMANA",f"{len(pag_sem)} conta(s) nos próximos 7 dias — Total: {fmt_brl(sum(fmt_v(p["valor"]) for p in pag_sem))}",COR_AMARELO_CLARO,"B7950B"))
        if rec_venc: alertas.append(("🔴 COBRANÇAS VENCIDAS",f"{len(rec_venc)} cobrança(s) vencida(s) — Total: {fmt_brl(sum(fmt_v(r["valor"]) for r in rec_venc))}",COR_VERMELHO_CLARO,COR_VERMELHO))
        if orc_pend: alertas.append(("🟡 ORÇAMENTOS PENDENTES",f"{len(orc_pend)} orçamento(s) aguardando — Total: {fmt_brl(sum(fmt_v(o["valor"]) for o in orc_pend))}",COR_AMARELO_CLARO,"B7950B"))
        if not alertas: alertas.append(("✅ TUDO EM DIA","Nenhuma pendência crítica.",COR_VERDE_CLARO,COR_VERDE))

        for i,(status,msg,bg_c,fg_c) in enumerate(alertas):
            r=row+1+i; ws.row_dimensions[r].height=20
            ws.merge_cells(f"A{r}:B{r}"); ws.merge_cells(f"C{r}:H{r}")
            c1=ws.cell(row=r,column=1,value=status)
            c1.font=fn(bold=True,size=9,cor=fg_c); c1.fill=fl(bg_c); c1.alignment=al("center","center"); c1.border=bd()
            c2=ws.cell(row=r,column=3,value=msg)
            c2.font=fn(size=9,cor="333333"); c2.fill=fl(bg_c); c2.alignment=al("left","center"); c2.border=bd()
            ws.cell(row=r,column=2).fill=fl(bg_c); ws.cell(row=r,column=2).border=bd()

        row_sum=row+len(alertas)+2; ws.row_dimensions[row_sum].height=22
        ws.merge_cells(f"A{row_sum}:H{row_sum}")
        c=ws.cell(row=row_sum,column=1,value="📈  RESUMO FINANCEIRO CONSOLIDADO")
        c.font=fn(bold=True,size=11,cor=COR_BRANCO); c.fill=fl(COR_AZUL); c.alignment=al("left","center")

        resumo_dados=[
            ("A PAGAR","Em aberto",len(pag_ab),total_pagar,COR_VERMELHO_CLARO),
            ("A PAGAR","Vencidas",len(pag_venc),sum(fmt_v(p["valor"]) for p in pag_venc),COR_VERMELHO_CLARO),
            ("A PAGAR","Vence hoje",len(pag_hoje),sum(fmt_v(p["valor"]) for p in pag_hoje),COR_LARANJA_CLARO),
            ("A PAGAR","Esta semana",len(pag_sem),sum(fmt_v(p["valor"]) for p in pag_sem),COR_AMARELO_CLARO),
            ("A PAGAR","Pagas",len(pag_pg),sum(fmt_v(p["valor"]) for p in pag_pg),COR_VERDE_CLARO),
            (None,None,None,None,COR_BRANCO),
            ("A RECEBER","Em aberto",len(rec_ab),total_receber,COR_AZUL_CLARO),
            ("A RECEBER","Vencidas",len(rec_venc),sum(fmt_v(r["valor"]) for r in rec_venc),COR_VERMELHO_CLARO),
            ("A RECEBER","Recebidas",len(rec_pg),total_recebido,COR_VERDE_CLARO),
            (None,None,None,None,COR_BRANCO),
            ("ORÇAMENTOS","Pendentes",len(orc_pend),total_orc_pend,COR_AMARELO_CLARO),
            ("ORÇAMENTOS","Aprovados",len(orc_aprov),sum(fmt_v(o["valor"]) for o in orc_aprov),COR_VERDE_CLARO),
            ("ORÇAMENTOS","Recusados",len(orc_rec),sum(fmt_v(o["valor"]) for o in orc_rec),COR_VERMELHO_CLARO),
        ]
        r_h=row_sum+1; ws.row_dimensions[r_h].height=18
        for col,h in enumerate(["MÓDULO","CATEGORIA","QTD","VALOR TOTAL"],1):
            c=ws.cell(row=r_h,column=col,value=h)
            c.font=fn(bold=True,size=9,cor=COR_BRANCO); c.fill=fl(COR_HEADER_DARK); c.alignment=al("center"); c.border=bd()
        for col in range(5,9): ws.cell(row=r_h,column=col).fill=fl(COR_HEADER_DARK); ws.cell(row=r_h,column=col).border=bd()
        ws.merge_cells(f"E{r_h}:H{r_h}")
        for i,(modulo,cat,qtd,valor,bg_c) in enumerate(resumo_dados):
            r=r_h+1+i; ws.row_dimensions[r].height=18
            if modulo is None:
                ws.row_dimensions[r].height=6
                for col in range(1,9): ws.cell(row=r,column=col).fill=fl(COR_BRANCO)
                continue
            ws.merge_cells(f"E{r}:H{r}")
            for col,val in enumerate([modulo,cat,qtd,valor],1):
                c=ws.cell(row=r,column=col,value=val)
                c.font=fn(size=9,bold=(col==1)); c.fill=fl(bg_c); c.border=bd()
                if col==3: c.alignment=al("center")
                if col==4: c.number_format="R$ #,##0.00"; c.alignment=al("right")
            ws.cell(row=r,column=5).fill=fl(bg_c); ws.cell(row=r,column=5).border=bd()

        ws.column_dimensions["A"].width=16; ws.column_dimensions["B"].width=22
        ws.column_dimensions["C"].width=8; ws.column_dimensions["D"].width=18
        for cc in ["E","F","G","H"]: ws.column_dimensions[cc].width=14

        # ── A PAGAR ──
        ws2=wb.create_sheet("💸 A Pagar"); ws2.sheet_view.showGridLines=False
        titulo_sheet(ws2,"  KEYCOMERCE  |  CONTAS A PAGAR",f"  Gerado em {data_str} às {hora_str}",COR_VERMELHO)
        merge_titulo(ws2,9,COR_VERMELHO)
        header_row(ws2,["ID","Fornecedor / Descrição","Valor (R$)","Vencimento","Tipo Pgto","Status","Dias","Data Pagamento","Observações"],row=4,bg=COR_VERMELHO)
        sf={"pago":(COR_VERDE_CLARO,COR_VERDE),"vencido":(COR_VERMELHO_CLARO,COR_VERMELHO),"hoje":(COR_LARANJA_CLARO,COR_LARANJA),"semana":(COR_AMARELO_CLARO,"B7950B"),"aberto":(COR_AZUL_CLARO,COR_AZUL)}
        for p in sorted(db["pagar"],key=lambda x:(0 if x["status"]=="aberto" else 1,days_until(x["vencimento"]))):
            d=days_until(p["vencimento"])
            if p["status"]=="pago": st="pago";st_txt="✅ Pago";dias_txt="—"
            elif d<0: st="vencido";st_txt="🔴 Vencida";dias_txt=f"{abs(d)}d atrás"
            elif d==0: st="hoje";st_txt="🔴 Vence Hoje";dias_txt="HOJE"
            elif d<=7: st="semana";st_txt="🟡 Esta Semana";dias_txt=f"Em {d}d"
            else: st="aberto";st_txt="🟢 Em Aberto";dias_txt=f"Em {d}d"
            bg_c,fg_c=sf[st]
            ws2.append([p["id"],p["fornecedor"],fmt_v(p["valor"]),p["vencimento"],p.get("tipo","—"),st_txt,dias_txt,p.get("data_pagamento","—"),p.get("obs","")])
            r=ws2.max_row; ws2.row_dimensions[r].height=18
            for col,cell in enumerate(ws2[r],1):
                cell.fill=fl(bg_c); cell.border=bd()
                cell.font=fn(size=9,bold=(col==1),cor=fg_c if col==6 else COR_PRETO)
                if col==3: cell.number_format="R$ #,##0.00"; cell.alignment=al("right")
                elif col in [4,7,8]: cell.alignment=al("center")
                else: cell.alignment=al("left")
        ws2.append([])
        for lbl,total in [("TOTAL EM ABERTO:",sum(fmt_v(p["valor"]) for p in pag_ab)),("TOTAL PAGO:",sum(fmt_v(p["valor"]) for p in pag_pg)),("TOTAL GERAL:",sum(fmt_v(p["valor"]) for p in db["pagar"]))]:
            ws2.append(["",lbl,total])
            r=ws2.max_row
            ws2.cell(row=r,column=2).font=fn(bold=True,size=9); ws2.cell(row=r,column=2).fill=fl(COR_CINZA_CLARO); ws2.cell(row=r,column=2).border=bd()
            c3=ws2.cell(row=r,column=3); c3.number_format="R$ #,##0.00"; c3.font=fn(bold=True,size=10,cor=COR_VERMELHO); c3.fill=fl(COR_CINZA_CLARO); c3.border=bd(); c3.alignment=al("right")
        ws2.column_dimensions["A"].width=10; ws2.column_dimensions["B"].width=35; ws2.column_dimensions["C"].width=16
        ws2.column_dimensions["D"].width=14; ws2.column_dimensions["E"].width=16; ws2.column_dimensions["F"].width=18
        ws2.column_dimensions["G"].width=14; ws2.column_dimensions["H"].width=18; ws2.column_dimensions["I"].width=35

        # ── A RECEBER ──
        ws3=wb.create_sheet("📥 A Receber"); ws3.sheet_view.showGridLines=False
        titulo_sheet(ws3,"  KEYCOMERCE  |  CONTAS A RECEBER",f"  Gerado em {data_str} às {hora_str}",COR_VERDE)
        merge_titulo(ws3,9,COR_VERDE)
        header_row(ws3,["ID","Cliente","Valor (R$)","Vencimento","Status","Dias","NF","Pedido","Observações"],row=4,bg=COR_VERDE)
        for r_item in sorted(db["receber"],key=lambda x:(0 if x["status"]=="aberto" else 1,days_until(x["vencimento"]))):
            d=days_until(r_item["vencimento"])
            if r_item["status"]=="pago": st="pago";st_txt="✅ Recebido";dias_txt="—"
            elif d<0: st="vencido";st_txt="🔴 Vencida";dias_txt=f"{abs(d)}d atrás"
            elif d==0: st="hoje";st_txt="🔴 Vence Hoje";dias_txt="HOJE"
            elif d<=7: st="semana";st_txt="🟡 Esta Semana";dias_txt=f"Em {d}d"
            else: st="aberto";st_txt="🟢 Em Aberto";dias_txt=f"Em {d}d"
            bg_c,fg_c=sf[st]
            ws3.append([r_item["id"],r_item["cliente"],fmt_v(r_item["valor"]),r_item["vencimento"],st_txt,dias_txt,r_item.get("nf","—"),r_item.get("pedido","—"),r_item.get("obs","")])
            r=ws3.max_row; ws3.row_dimensions[r].height=18
            for col,cell in enumerate(ws3[r],1):
                cell.fill=fl(bg_c); cell.border=bd()
                cell.font=fn(size=9,bold=(col==1),cor=fg_c if col==5 else COR_PRETO)
                if col==3: cell.number_format="R$ #,##0.00"; cell.alignment=al("right")
                elif col in [4,6,7,8]: cell.alignment=al("center")
                else: cell.alignment=al("left")
        ws3.append([])
        for lbl,total in [("TOTAL EM ABERTO:",sum(fmt_v(r["valor"]) for r in rec_ab)),("TOTAL RECEBIDO:",sum(fmt_v(r["valor"]) for r in rec_pg)),("TOTAL GERAL:",sum(fmt_v(r["valor"]) for r in db["receber"]))]:
            ws3.append(["",lbl,total])
            r=ws3.max_row
            ws3.cell(row=r,column=2).font=fn(bold=True,size=9); ws3.cell(row=r,column=2).fill=fl(COR_CINZA_CLARO); ws3.cell(row=r,column=2).border=bd()
            c3=ws3.cell(row=r,column=3); c3.number_format="R$ #,##0.00"; c3.font=fn(bold=True,size=10,cor=COR_VERDE); c3.fill=fl(COR_CINZA_CLARO); c3.border=bd(); c3.alignment=al("right")
        ws3.column_dimensions["A"].width=10; ws3.column_dimensions["B"].width=30; ws3.column_dimensions["C"].width=16
        ws3.column_dimensions["D"].width=14; ws3.column_dimensions["E"].width=18; ws3.column_dimensions["F"].width=14
        ws3.column_dimensions["G"].width=14; ws3.column_dimensions["H"].width=12; ws3.column_dimensions["I"].width=35

        # ── ORÇAMENTOS ──
        ws4=wb.create_sheet("📋 Orçamentos"); ws4.sheet_view.showGridLines=False
        titulo_sheet(ws4,"  KEYCOMERCE  |  ORÇAMENTOS",f"  Gerado em {data_str} às {hora_str}",COR_LARANJA)
        merge_titulo(ws4,8,COR_LARANJA)
        header_row(ws4,["ID","Cliente","Descrição","Valor (R$)","Tipo","Canal","Status","Data"],row=4,bg=COR_LARANJA)
        osf={"pendente":(COR_AMARELO_CLARO,"B7950B"),"aprovado":(COR_VERDE_CLARO,COR_VERDE),"recusado":(COR_VERMELHO_CLARO,COR_VERMELHO),"enviado":(COR_AZUL_CLARO,COR_AZUL)}
        ost={"pendente":"🟡 Pendente","aprovado":"🟢 Aprovado","recusado":"🔴 Recusado","enviado":"🔵 Enviado"}
        for o in db["orcamentos"]:
            bg_c,fg_c=osf.get(o["status"],(COR_CINZA_CLARO,COR_CINZA))
            ws4.append([o["id"],o["cliente"],o["descricao"],fmt_v(o["valor"]),o.get("tipo","—"),o.get("canal","—"),ost.get(o["status"],o["status"]),o["data"]])
            r=ws4.max_row; ws4.row_dimensions[r].height=18
            for col,cell in enumerate(ws4[r],1):
                cell.fill=fl(bg_c); cell.border=bd()
                cell.font=fn(size=9,bold=(col==1),cor=fg_c if col==7 else COR_PRETO)
                if col==4: cell.number_format="R$ #,##0.00"; cell.alignment=al("right")
                elif col in [6,7,8]: cell.alignment=al("center")
                else: cell.alignment=al("left")
        ws4.append([])
        for lbl,total in [("PENDENTES:",sum(fmt_v(o["valor"]) for o in orc_pend)),("APROVADOS:",sum(fmt_v(o["valor"]) for o in orc_aprov)),("TOTAL GERAL:",sum(fmt_v(o["valor"]) for o in db["orcamentos"]))]:
            ws4.append(["",lbl,total])
            r=ws4.max_row
            ws4.cell(row=r,column=2).font=fn(bold=True,size=9); ws4.cell(row=r,column=2).fill=fl(COR_CINZA_CLARO); ws4.cell(row=r,column=2).border=bd()
            c3=ws4.cell(row=r,column=3); c3.number_format="R$ #,##0.00"; c3.font=fn(bold=True,size=10,cor=COR_LARANJA); c3.fill=fl(COR_CINZA_CLARO); c3.border=bd(); c3.alignment=al("right")
        ws4.column_dimensions["A"].width=10; ws4.column_dimensions["B"].width=28; ws4.column_dimensions["C"].width=40
        ws4.column_dimensions["D"].width=16; ws4.column_dimensions["E"].width=22; ws4.column_dimensions["F"].width=14
        ws4.column_dimensions["G"].width=16; ws4.column_dimensions["H"].width=14

        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        return buf.getvalue()
    except Exception as e:
        print(f"[EXCEL] Erro gerar: {e}")
        import traceback; traceback.print_exc()
        return None


def enviar_planilha(chat_id, motivo="📊 Planilha diária"):
    if not EXCEL_OK:
        send_message(chat_id, "❌ openpyxl não instalado.")
        return
    excel_bytes = gerar_excel()
    if not excel_bytes:
        send_message(chat_id, "❌ Erro ao gerar planilha.")
        return
    try:
        nome = f"Keycomerce_{datetime.now().strftime('%d%m%Y')}.xlsx"
        boundary = "----KeycoBoundary"
        body = (
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="chat_id"\r\n\r\n{chat_id}\r\n'
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="caption"\r\n\r\n{motivo} — {datetime.now().strftime("%d/%m/%Y")}\r\n'
            f"--{boundary}\r\n"
            f'Content-Disposition: form-data; name="document"; filename="{nome}"\r\n'
            f"Content-Type: application/vnd.openxmlformats-officedocument.spreadsheetml.sheet\r\n\r\n"
        ).encode() + excel_bytes + f"\r\n--{boundary}--\r\n".encode()

        req = urllib.request.Request(
            f"{TELEGRAM_API}/sendDocument",
            data=body,
            headers={"Content-Type": f"multipart/form-data; boundary={boundary}"}
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read())
            if result.get("ok"):
                print(f"[EXCEL] Planilha enviada para {chat_id}")
            else:
                print(f"[EXCEL] Erro Telegram: {result}")
    except Exception as e:
        print(f"[EXCEL] Erro envio: {e}")
        send_message(chat_id, f"❌ Erro ao enviar planilha: {e}")

def enviar_email_planilha(excel_bytes, data_str):
    try:
        import base64 as b64mod
        excel_b64 = b64mod.b64encode(excel_bytes).decode()
        nome_arquivo = f"Keycomerce_{data_str.replace('/','')}.xlsx"
        payload = json.dumps({
            "from": "Keycomerce <onboarding@resend.dev>",
            "to": [EMAIL_TO],
            "subject": f"Keycomerce — Planilha Financeira {data_str}",
            "html": f"<h2>Keycomerce Gestão</h2><p>Bom dia!</p><p>Segue em anexo a planilha financeira da Keyco referente a <strong>{data_str}</strong>.</p><p>Keycomerce Bot</p>",
            "attachments": [{
                "filename": nome_arquivo,
                "content": excel_b64
            }]
        }).encode()
        req = urllib.request.Request(
            "https://api.resend.com/emails",
            data=payload,
            headers={
                "Authorization": f"Bearer {RESEND_KEY}",
                "Content-Type": "application/json"
            }
        )
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read())
            print(f"[EMAIL] Enviado: {result}")
            return True
    except Exception as e:
        print(f"[EMAIL] Erro: {e}")
        return False

def enviar_alerta_diario(chat_id):
    db = load_data()
    now = datetime.now()
    dias = ["Segunda","Terca","Quarta","Quinta","Sexta","Sabado","Domingo"]
    dia = dias[now.weekday()]
    data_fmt = now.strftime("%d/%m/%Y")

    pag_ab = [p for p in db["pagar"] if p["status"]=="aberto"]
    pag_venc = [p for p in pag_ab if days_until(p["vencimento"])<0]
    pag_hoje = [p for p in pag_ab if days_until(p["vencimento"])==0]
    pag_sem = [p for p in pag_ab if 1<=days_until(p["vencimento"])<=7]
    rec_ab = [r for r in db["receber"] if r["status"]=="aberto"]
    rec_venc = [r for r in rec_ab if days_until(r["vencimento"])<0]
    rec_hoje = [r for r in rec_ab if days_until(r["vencimento"])==0]

    msg = "Bom dia! " + dia + ", " + data_fmt + "\n\n"

    if pag_venc:
        msg += "🔴 *CONTAS VENCIDAS:*\n"
        for p in pag_venc:
            d = abs(days_until(p["vencimento"]))
            msg += "  • " + p["id"] + " — " + p["fornecedor"] + "\n"
            msg += "    " + fmt_brl(p["valor"]) + " · venceu ha " + str(d) + "d\n"
        msg += "\n"

    if pag_hoje:
        msg += "🔴 *VENCE HOJE:*\n"
        for p in pag_hoje:
            msg += "  • " + p["id"] + " — " + p["fornecedor"] + "\n"
            msg += "    " + fmt_brl(p["valor"]) + "\n"
        msg += "\n"

    if pag_sem:
        msg += "🟡 *VENCE ESTA SEMANA:*\n"
        for p in sorted(pag_sem, key=lambda x: days_until(x["vencimento"])):
            d = days_until(p["vencimento"])
            msg += "  • " + p["id"] + " — " + p["fornecedor"] + "\n"
            msg += "    " + fmt_brl(p["valor"]) + " · em " + str(d) + "d (" + p["vencimento"] + ")\n"
        msg += "\n"

    if rec_venc or rec_hoje:
        msg += "📥 *COBRANÇAS VENCIDAS/HOJE:*\n"
        for r in rec_venc:
            d = abs(days_until(r["vencimento"]))
            msg += "  • " + r["id"] + " — " + r["cliente"] + "\n"
            msg += "    " + fmt_brl(r["valor"]) + " · venceu ha " + str(d) + "d\n"
        for r in rec_hoje:
            msg += "  • " + r["id"] + " — " + r["cliente"] + "\n"
            msg += "    " + fmt_brl(r["valor"]) + " · vence hoje\n"
        msg += "\n"

    urgente = sum(float(p["valor"]) for p in pag_venc+pag_hoje)
    semana_total = sum(float(p["valor"]) for p in pag_sem)
    a_receber = sum(float(r["valor"]) for r in rec_ab)

    msg += "─────────────────\n"
    if urgente > 0:
        msg += "🔴 *Urgente pagar:* " + fmt_brl(urgente) + "\n"
    if semana_total > 0:
        msg += "🟡 *Semana:* " + fmt_brl(semana_total) + "\n"
    msg += "📥 *A receber:* " + fmt_brl(a_receber) + "\n"

    if not pag_venc and not pag_hoje and not pag_sem and not rec_venc and not rec_hoje:
        msg = "Bom dia! " + dia + ", " + data_fmt + "\n\n✅ Nenhuma pendencia critica hoje. Bom dia!"

    send_message(chat_id, msg)

def setup_webhook():
    domain = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "")
    if domain:
        url = f"https://{domain}/webhook"
        try:
            http_post(f"{TELEGRAM_API}/setWebhook", {"url": url})
            print(f"Webhook: {url}")
        except Exception as e:
            print(f"Erro webhook: {e}")

# ── WEBHOOK ──
class WebhookHandler(BaseHTTPRequestHandler):
    def do_POST(self):
        if self.path == "/webhook":
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length)
            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"ok")
            try:
                update = json.loads(body)
                if "callback_query" in update:
                    cb = update["callback_query"]
                    chat_id = cb["message"]["chat"]["id"]
                    data_cb = cb["data"]
                    answer_callback(cb["id"])
                    if data_cb.startswith("confirmar_boleto_"):
                        tipo = data_cb.replace("confirmar_boleto_","")
                        state = user_states.get(chat_id, {})
                        temp = state.get("temp", {})
                        db = load_data()
                        if tipo == "pagar":
                            item = {"id": next_id(db,"PAG"), "fornecedor": temp.get("fornecedor","Sem nome"), "valor": float(temp.get("valor",0)), "vencimento": temp.get("vencimento",""), "status": "aberto", "tipo": "Boleto"}
                            db["pagar"].append(item)
                            save_data(db)
                            user_states[chat_id] = {}
                            threading.Thread(target=send_message, args=(chat_id, "Lancado " + item["id"] + " — " + item["fornecedor"] + " " + fmt_brl(item["valor"]), [[{"text": "💸 Ver contas", "callback_data": "menu_pagar"}, {"text": "🏠 Menu", "callback_data": "menu_principal"}]])).start()
                        else:
                            item = {"id": next_id(db,"COB"), "cliente": temp.get("fornecedor","Sem nome"), "valor": float(temp.get("valor",0)), "vencimento": temp.get("vencimento",""), "status": "aberto", "nf": ""}
                            db["receber"].append(item)
                            save_data(db)
                            user_states[chat_id] = {}
                            threading.Thread(target=send_message, args=(chat_id, "Lancado " + item["id"], [[{"text": "🏠 Menu", "callback_data": "menu_principal"}]])).start()
                    elif data_cb.startswith("orc_aprovar_"):
                        orc_id = data_cb.replace("orc_aprovar_","")
                        db = load_data()
                        orc = next((o for o in db["orcamentos"] if o["id"]==orc_id), None)
                        if orc:
                            orc["status"] = "aprovado"
                            save_data(db)
                            threading.Thread(target=send_message, args=(chat_id, "Orcamento " + orc["id"] + " aprovado!", [[{"text": "📋 Orçamentos", "callback_data": "menu_orcamentos"}]])).start()
                    elif data_cb.startswith("orc_recusar_"):
                        orc_id = data_cb.replace("orc_recusar_","")
                        db = load_data()
                        orc = next((o for o in db["orcamentos"] if o["id"]==orc_id), None)
                        if orc:
                            orc["status"] = "recusado"
                            save_data(db)
                            threading.Thread(target=send_message, args=(chat_id, "Orcamento " + orc["id"] + " recusado.", [[{"text": "📋 Orçamentos", "callback_data": "menu_orcamentos"}]])).start()
                    else:
                        threading.Thread(target=handle_callback, args=(chat_id, data_cb)).start()
                elif "message" in update:
                    message = update["message"]
                    chat_id = message.get("chat", {}).get("id")
                    if chat_id:
                        text = message.get("text")
                        photo = message.get("photo")
                        document = message.get("document")
                        threading.Thread(target=handle_message, args=(chat_id, text, photo, document)).start()
            except Exception as e:
                print(f"Erro webhook: {e}")
        else:
            self.send_response(200)
            self.end_headers()
            self.wfile.write(b"ok")

    def do_GET(self):
        self.send_response(200)
        self.end_headers()
        self.wfile.write(b"Keycomerce Bot OK")

    def log_message(self, format, *args):
        pass


if __name__ == "__main__":
    if EXCEL_OK:
        scheduler_thread = threading.Thread(target=scheduler, daemon=True)
        scheduler_thread.start()
        print("[SCHEDULER] Thread iniciada")
    port = int(os.environ.get("PORT", 8080))
    setup_webhook()
    print(f"Keycomerce Bot rodando na porta {port}")
    server = HTTPServer(("0.0.0.0", port), WebhookHandler)
    server.serve_forever()
